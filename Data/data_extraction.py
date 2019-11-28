import os, sys
from collections import OrderedDict
from openmesh import *
from openpyxl import load_workbook, workbook
import numpy as np
from scipy.io import loadmat
from pprint import pprint

from pymongo import MongoClient, errors as mongo_error
from pprint import pprint
from copy import deepcopy
import itertools
import tqdm
import h5py


class DataExtractor:
    def __init__(self, path, mongo_client, db="OptiDrape", db_collection="training", update_db=True):

        self._path = path
        self._mesh_format = '.pc'
        self._shear_format = '.asc'
        self._db = None
        self._update_db = update_db

        self._mongo_client = mongo_client
        self._db = self._mongo_client[db]

        self._general_format = '.asc'

        self._file_dict = dict()

        self._mesh_dict = {}

        self._meshes = OrderedDict()

    def set_file_format(self, mesh=None, shear=None):
        if mesh:
            self._mesh_format = mesh

        if shear:
            self._shear_format = shear

    def load_file_from_esi(self):
        self._file_dict = {'mesh': {}, 'shear': {}}
        try:
            _all_file = os.listdir(self._path)
            _file_counter = [0, 0]
            for _file in _all_file:
                _split = os.path.splitext(_file)
                if _split[1] == self._mesh_format:
                    self._file_dict['mesh'].update({_split[0]: _file})
                    _file_counter[0] += 1
                elif _split[1] == self._shear_format:
                    self._file_dict['shear'].update({_split[0]: _file})
                    _file_counter[1] += 1

            if _file_counter[0] is 0:
                print("Warning: There was no mesh file found.")

            if _file_counter[1] is 0:
                print("Warning: There was no shear file found.")
        except FileNotFoundError:
            print("Error: Invalid path!")

    def load_and_extract_file_from_esi(self):
        # because each folder do has sub folder, so make a iteration of all of them
        def extract_point(name, point, value):
            if 'X' in name.upper():
                point[0] = float(value)
            elif 'Y' in name.upper():
                point[1] = float(value)
            elif 'Z' in name.upper():
                point[2] = float(value)
            return point

        def extract_content(name, value, default=None):
            if 'Scherwinkel' in name:
                return {'shear_angle': float(value)}
            elif 'Thinning' in name:
                return {'thining': float(value)}
            elif 'Pos' in name and '=' in name:  # this indicates the position at preform begin
                if not default or 'point_start' not in default.keys():
                    point = [0, 0, 0]
                    return {'point_start': extract_point(name, point, value)}
                else:
                    _point = default['point_start']
                    return {'point_start': extract_point(name, _point, value)}
            elif 'Pos' in name:  # this indicates the position at preform end
                if not default or 'point_end' not in default.keys():
                    point = [0, 0, 0]
                    return {'point_end': extract_point(name, point, value)}
                else:
                    _point = default['point_end']
                    return {'point_end': extract_point(name, _point, value)}

        for d in os.walk(self._path):
            # pprint(d)
            # format of d: [path, subfolder name, file names]
            if d[2]:
                # if the third list is not null, then we could begin to iterate the files
                ## sub_dirs: a list of names of the path,
                ### first one is the form name
                ### second one is the
                _sub_dirs = [s for s in d[0].split("/") if s not in self._path.split("/")]
                # pprint(_sub_dirs)
                if _sub_dirs[0] not in self._mesh_dict.keys():
                    self._mesh_dict.update({_sub_dirs[0]: dict()})
                if _sub_dirs[1] not in self._mesh_dict[_sub_dirs[0]].keys():
                    self._mesh_dict[_sub_dirs[0]].update({_sub_dirs[1]: dict()})
                if _sub_dirs[2] not in self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]].keys():
                    self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]].update({_sub_dirs[2]: dict()})

                self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]].update({'Vertex': dict()})
                for file in d[2]:
                    if '.asc' in file:
                        _began_read_line = False
                        with open(d[0] + '/' + file, 'r') as this_file:
                            _lines = this_file.readlines()
                            _filename_to_extract = file.replace('.asc', '')
                            for line in _lines:
                                _line = [s for s in line.split(' ') if s is not '']
                                if _began_read_line:
                                    try:
                                        if _line[0] not in self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]][
                                            'Vertex'].keys():
                                            # print(_line)
                                            self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]]['Vertex'].update(
                                                {_line[0]:
                                                     extract_content(name=_filename_to_extract, value=_line[1])
                                                 })
                                        else:
                                            _this_dict = \
                                            self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]]['Vertex'][
                                                _line[0]]
                                            self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]]['Vertex'][
                                                _line[0]].update(
                                                extract_content(name=_filename_to_extract, value=_line[1],
                                                                default=_this_dict))
                                    except IndexError:
                                        pass

                                if all(s in _line for s in ['Number']):
                                    _began_read_line = True

                # _this_mesh_normal = TriMesh()
                # _this_mesh_T0 = TriMesh
                # _this_mesh_normal.request_vertex_normals()
                # _this_mesh_normal.request_vertex_texcoords2D()
                # _this_mesh_normal.request_face_normals()
                # _this_mesh_normal.request_vertex_colors()

                # for id in self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]].keys():
                # pprint(_sub_dirs[0]+_sub_dirs[1]+_sub_dirs[2]+str(id))
                # pprint(self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]][id])

                # _point = self._mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]][id]['point_start']
                # _this_mesh_normal.add_vertex(TriMesh.Point(_point[0], _point[1], _point[2]))

                # self._meshes.update({_sub_dirs[0]+'.'+_sub_dirs[1]+'.'+_sub_dirs[2]+'.normal': _this_mesh_normal})
        for geometry_name in self._mesh_dict.keys():
            if self._update_db:
                for geometry in self._mesh_dict[geometry_name].keys():
                    self._db[geometry_name].insert_one({geometry: self._mesh_dict[geometry_name][geometry]})

    def extract_mesh_from_matlab(self):

        pass

    def extract_mesh_from_esi(self):
        for mesh_file in self._file_dict['mesh'].keys():
            _lines = []

            _node_dict = OrderedDict()
            _shell_dict = OrderedDict()

            with open(self._path + self._file_dict['mesh'][mesh_file]) as mesh:
                _lines = mesh.readlines()
            for _line in _lines:
                _split_list = [s for s in _line.split(' ') if s is not '']

                # extract nodes
                if _split_list[0] == 'NODE':
                    _node_dict.update({_split_list[2]: _split_list[3:6]})

                elif _split_list[0] == 'SHELL':
                    _shell_dict.update({_split_list[2]: _split_list[4:8]})

            self._mesh_dict.update({mesh_file: {'node': _node_dict,
                                                'shell': _shell_dict}})

    def rebuild_mesh_from_esi(self):
        for mesh_key in self._mesh_dict.keys():
            _this_mesh = TriMesh()
            _this_mesh.request_vertex_normals()
            _this_mesh.request_vertex_texcoords2D()
            _this_mesh.request_face_normals()
            _this_mesh.request_vertex_colors()
            _node_dict = {}
            for node_key in self._mesh_dict[mesh_key]['node'].keys():
                _node = [float(s) for s in self._mesh_dict[mesh_key]['node'][node_key]]

                _node_dict.update({node_key: _this_mesh.add_vertex(PolyMesh.Point(_node[0], _node[1], _node[2]))})
            _shell_dict = {}
            for shell_key in self._mesh_dict[mesh_key]['shell'].keys():
                _shell_consist = self._mesh_dict[mesh_key]['shell'][shell_key]
                _shell_dict.update({shell_key: _this_mesh.add_face([_node_dict[node] for node in _shell_consist])})

            self._meshes.update({mesh_key: {'mesh': _this_mesh,
                                            'node': _node_dict,
                                            'shell': _shell_dict}})

            _this_mesh.update_vertex_normals()
            _this_mesh.update_face_normals()

            for fh in _this_mesh.faces():
                for vh in _this_mesh.fv(fh):
                    print(_this_mesh.color(vh)[1])
                # print(_this_mesh.normal(fh)[0], _this_mesh.normal(fh)[1], _this_mesh.normal(fh)[2])

                # break

    def rebuild_mesh_from_esi_2(self):

        pass

    def save_meshes(self):
        print("All Meshes extracted will be saved in the same Path.")
        for mesh_key in self._meshes.keys():
            print(write_mesh(self._meshes[mesh_key]['mesh'], self._path + 'atest' + '.obj'))

    def get_mesh(self):
        return self._meshes

    def get_mesh_dict(self):
        return self._mesh_dict

    def test_write_mesh(self):
        print(write_mesh(self._meshes['Elipsenstumpf']['E150_150']['L_3_55_4545']['normal'],
                         '/home/haoming/Desktop/Data/Elipsenstumpf' + '.obj'))


def load_and_extract_data_from_simulation(path, db_instance=None):
    # because each folder do has sub folder, so make a iteration of all of them
    def extract_point(name, point, value):
        if 'X' in name.upper():
            point[0] = float(value)
        elif 'Z' in name.upper():
            point[1] = float(value)
        elif 'Y' in name.upper():
            point[2] = float(value)
        return point

    def extract_content(name, value, default=None):
        if 'Scherwinkel' in name:
            return {'shear_angle': float(value)}
        elif 'Thinning' in name:
            return {'thinning': float(value)}
        elif 'Pos' in name and '=' in name:  # this indicates the position at preform begin
            if not default or 'point_start' not in default.keys():
                point = [0, 0, 0]
                return {'point_start': extract_point(name, point, value)}
            else:
                _point = default['point_start']
                return {'point_start': extract_point(name, _point, value)}
        elif 'Pos' in name:  # this indicates the position at preform end
            if not default or 'point_end' not in default.keys():
                point = [0, 0, 0]
                return {'point_end': extract_point(name, point, value)}
            else:
                _point = default['point_end']
                return {'point_end': extract_point(name, _point, value)}

    _mesh_dict = dict()
    _textile_dict = dict()
    for d in os.walk(path):
        # format of d: [path, subfolder name, file names]
        if d[2]:
            # if the third list is not null, then we could begin to iterate the files
            ## sub_dirs: a list of names of the path,
            ### first one is the form name
            ### second one is the
            _sub_dirs = [s for s in d[0].split("/") if s not in path.split("/")]
            # pprint(_sub_dirs)
            if _sub_dirs[0] not in _mesh_dict.keys():
                # _sub_dirs[0] == Name of Bauteil
                _mesh_dict.update({_sub_dirs[0]: dict()})
            if _sub_dirs[1] not in _mesh_dict[_sub_dirs[0]].keys():
                # _sub_dirs[1] == Name of Baulteilgeometrie
                _mesh_dict[_sub_dirs[0]].update({_sub_dirs[1]: dict()})

            if _sub_dirs[2] not in _mesh_dict[_sub_dirs[0]][_sub_dirs[1]].keys():
                # _sub_dirs[2] == Name of Textil
                _mesh_dict[_sub_dirs[0]][_sub_dirs[1]].update({_sub_dirs[2]: dict()})

                # TODO: we want to split the textil form from textil name
                ## B412C_4545 to B412C: 4545 and 090
                _textile_dict.update({_sub_dirs[2]: dict()})

            _mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]].update({'Vertex': dict()})
            # then we iterate all of the dirs to grasp the data
            for file in d[2]:
                if '.asc' in file:
                    _began_read_line = False
                    try:
                        with open(d[0] + '/' + file, 'r') as this_file:
                            _lines = this_file.readlines()
                            _filename_to_extract = file.replace('.asc', '')
                            for line in _lines:
                                _line = [s for s in line.split(' ') if s is not '']
                                if _began_read_line:
                                    try:
                                        if _line[0] not in _mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]][
                                            'Vertex'].keys():
                                            _mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]]['Vertex'].update(
                                                {_line[0]:
                                                     extract_content(name=_filename_to_extract, value=_line[1])
                                                 })
                                        else:
                                            _this_dict = _mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]]['Vertex'][
                                                _line[0]]
                                            _mesh_dict[_sub_dirs[0]][_sub_dirs[1]][_sub_dirs[2]]['Vertex'][
                                                _line[0]].update(
                                                extract_content(name=_filename_to_extract, value=_line[1],
                                                                default=_this_dict))
                                    except IndexError:
                                        pass

                                if all(s in _line for s in ['Number']):
                                    _began_read_line = True
                    except IOError:
                        pass

    # we want to redefine the id of the vertex from 0

    for geometry in _mesh_dict.keys():
        for geometry_type in _mesh_dict[geometry].keys():
            for textile in _mesh_dict[geometry][geometry_type].keys():
                _min_id = 99999999
                for id in _mesh_dict[geometry][geometry_type][textile]['Vertex'].keys():
                    id_int = int(id)
                    if id_int < _min_id:
                        _min_id = id_int
                _new_vertex_dict = dict()

                for id in _mesh_dict[geometry][geometry_type][textile]['Vertex'].keys():
                    _new_vertex_dict.update(
                        {str(int(id) - _min_id): _mesh_dict[geometry][geometry_type][textile]['Vertex'][id]})

                _mesh_dict[geometry][geometry_type][textile]['Vertex'] = _new_vertex_dict

    for textil in _textile_dict.keys():
        for geometry in _mesh_dict.keys():
            for geometry_type in _mesh_dict[geometry].keys():
                for _textil in _mesh_dict[geometry][geometry_type].keys():
                    if _textil == textil:
                        if geometry not in _textile_dict[textil].keys():
                            _textile_dict[textil].update(
                                {geometry: {geometry_type: _mesh_dict[geometry][geometry_type][_textil]}})
                        else:
                            _textile_dict[textil][geometry].update(
                                {geometry_type: _mesh_dict[geometry][geometry_type][_textil]})

    if db_instance:
        """
                for geometry_name in self._mesh_dict.keys():
                for geometry in self._mesh_dict[geometry_name].keys():
                    self._db[geometry_name].insert_one({geometry: self._mesh_dict[geometry_name][geometry]})
        """

        for textil in _textile_dict.keys():
            for geometry in _textile_dict[textil].keys():
                db_instance[textil].insert_one({geometry: _textile_dict[textil][geometry]})

    return _textile_dict, _mesh_dict



def extract_mesh_from_matlab(path,
                             default_data_set='ZT',
                             extract_info={'origin': ['XG', 'YG', 'ZG'], 'result': ['XT', 'YT', 'ZT'],
                                           'shear_angle': 'Scherwinkel'},
                             dim_info={},
                             db_instance=None, info_cb=None):
    # because of the co-work, the extract of data will be separat, once a type of geometry, so the path must be a file path

    if not os.path.isfile(path):
        if info_cb:
            info_cb("Error", "Invalid file path %s given." % str(path))
        print("Invalid File Path %s given" % str(path))

    _geometry_dict = dict()

    try:
        _mat = h5py.File(path, 'r')
        if "pyramid" in str(path).lower():
            # here we start to extract the pyramiden
            # iteration order: height, width, angle
            # first we get the dimension of the geometry
            try:
                _min_height = dim_info['min_height']
                _max_height = dim_info['max_height']
                _min_width = dim_info['min_width']
                _max_width = dim_info['max_width']
                _min_angle = dim_info['min_angle']
                _max_angle = dim_info['max_angle']

                # before we start, check the data and the dim info

                # here begins the points collection
                _data = _mat[default_data_set]  # ndarray of z info, which contains matrixs for all geometries

                _dim_height = _data.shape[2]
                _dim_width = _data.shape[1]
                _dim_angle = _data.shape[0]

                _step_height = int((_max_height - _min_height) / (_dim_height - 1))
                _step_width = int((_max_width - _min_width) / (_dim_width - 1))
                _step_angle = int((_max_angle - _min_angle) / (_dim_angle - 1))

                if not (_max_height - _min_height) % _dim_height or \
                        not (_max_width - _min_width) % _dim_width or \
                        not (_max_angle - _min_angle) % _dim_angle:
                    if info_cb:
                        info_cb("Warning", "Size of Geometry dosenot fit the min/max.")
                    print("Warning: Size of Geometry dosenot fit the min/max.")

                with tqdm.tqdm(total=_dim_height * _dim_width * _dim_angle, desc="Extracting Geometry Pyramidenstumpf ..." ) as pbar:
                    for i_angle, i_height, i_width in itertools.product(range(_dim_angle), range(_dim_height), range(_dim_width)):
                        _ori_z_matrix = _mat[_mat['ZG'][i_angle, i_width, i_height]]
                        _z_matrix = _mat[_data[i_angle, i_width, i_height]]
                        if _z_matrix.size > 2:  # if the size of the geometry exists, then we could extract
                            _origin = []
                            _sim_result = []

                            for i, j in itertools.product(range(_z_matrix.shape[0]), range(_z_matrix.shape[1])):
                                _origin.append([i, j, _ori_z_matrix[i, j]])
                                _sim_result.append([i, j, _z_matrix[i, j]])

                            _this_geo_name = 'H' + str(_min_height + i_height * _step_height) + \
                                             'W' + str(_min_width + i_width * _step_width) + \
                                             'A' + str(_min_angle + i_angle * _step_angle)

                            _geometry_dict.update({_this_geo_name: {'origin': _ori_z_matrix,
                                                                    'result': _sim_result,
                                                                    'shear_angle':  _mat[_mat['Scherwinkel'][i_angle, i_width, i_height]]}})
                        pbar.update()
                return _geometry_dict

            except KeyError:
                if info_cb:
                    info_cb("Warning", "Invalid dimension info, please check!")
                pass

        elif "halbkugel" in str(path).lower():

            _min_height = dim_info['min_height']
            _max_height = dim_info['max_height']
            _min_radius = dim_info['min_radius']
            _max_radius = dim_info['max_radius']

            _dim_height = _mat['ZT'][0, :].size
            _dim_radius = _mat['ZT'][:, 0].size

            _step_height = int((_max_height - _min_height) / (_dim_height - 1))
            _step_radius = int((_max_radius - _min_radius) / (_dim_radius - 1))

            if not (_max_height - _min_height) % _dim_height or \
                    not (_max_radius - _min_radius) % _dim_radius:
                if info_cb:
                    info_cb("Warning", "Size of Geometry dosenot fit the min/max.")
                print("Warning: Size of Geometry dosenot fit the min/max.")

            with tqdm.tqdm(total=_mat['ZT'].size, desc="Extracting Geometry Halbkugel ..." ) as pbar:
                for i_radius, i_height in itertools.product(range(_dim_radius), range(_dim_height)):
                    _ori_z_matrix = _mat['ZG'][i_height, i_height]
                    _z_matrix = _mat['ZT'][i_height, i_height]

                    if _z_matrix.size:
                        _origin = []
                        _sim_result = []

                        for i, j in itertools.product(range(_z_matrix.shape[0]), range(_z_matrix.shape[1])):
                            _origin.append([i, j, _ori_z_matrix[i, j]])
                            _sim_result.append([i, j, _z_matrix[i, j]])

                        _this_geo_name = 'R' + str(_min_radius + i_radius * _step_radius) + \
                                         'H' + str(_min_height + i_height * _step_height)

                        _geometry_dict.update({_this_geo_name: {'origin': _ori_z_matrix,
                                                                'result': _sim_result,
                                                                'shearangle': _mat['Scherwinkel'][
                                                                    i_radius, i_height]}})
                    pbar.update(n=i_radius * i_height)

    except IOError:
        pass

        for content in extract_info:
            try:

                pass


            except ValueError:
                print("ValueError!")
                pass

            pass

    pass


def extract_meta_data(path, requests):
    if not path[0]:
        print('no data choosen')
        return

    _wb = load_workbook(path[0])
    _meta_data_dict = {}

    for _sheet_name in _wb.sheetnames:
        _current_sheet = _wb[_sheet_name]
        _data_package = {}

        for _row in _current_sheet.iter_rows(min_row=1, min_col=1, max_row=_current_sheet.max_row,
                                             max_col=_current_sheet.max_column):

            # TODO: check what kind of data should we extract
            for cell in _row:
                for _single_request in requests:
                    if _single_request in str(cell.value):
                        _data_package[_single_request] = _current_sheet.cell(row=cell.row,
                                                                             column=cell.col_idx + 1).value

        _meta_data_dict[_current_sheet.title] = _data_package

    print(_meta_data_dict)

    return _meta_data_dict


if __name__ == '__main__':
    test = DataExtractor(path='/home/haoming/Desktop/Data/new_data')
    test.load_and_extract_file_from_esi()
